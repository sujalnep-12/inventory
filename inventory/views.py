from datetime import timedelta
from decimal import Decimal

from django.contrib import messages
from django.db import transaction
from django.db.models import DecimalField, ExpressionWrapper, F, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import (
    DateRangeForm,
    ExpenseCategoryForm,
    ExpenseForm,
    ProductCategoryForm,
    ProductForm,
    SaleForm,
    SaleItemFormSet,
    ServiceCategoryForm,
    ServiceForm,
    StockAdjustmentForm,
)
from .models import (
    Expense,
    ExpenseCategory,
    Product,
    ProductCategory,
    Sale,
    SaleItem,
    Service,
    ServiceCategory,
    StockAdjustment,
)


ZERO = Decimal('0.00')


def money_sum(queryset, field_name):
    return queryset.aggregate(total=Coalesce(Sum(field_name), ZERO, output_field=DecimalField()))['total']


def selected_day(request):
    selected = parse_date(request.GET.get('date', ''))
    return selected or timezone.localdate()


def date_range_from_request(request):
    form = DateRangeForm(request.GET or None)
    today = timezone.localdate()
    start_date = today.replace(day=1)
    end_date = today
    if form.is_valid():
        start_date = form.cleaned_data.get('start_date') or start_date
        end_date = form.cleaned_data.get('end_date') or end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return form, start_date, end_date


def inventory_value():
    expression = ExpressionWrapper(
        F('stock_quantity') * F('cost_price'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    return Product.objects.aggregate(total=Coalesce(Sum(expression), ZERO, output_field=DecimalField()))['total']


def dashboard(request):
    day = selected_day(request)
    sales = Sale.objects.filter(sale_date=day)
    expenses = Expense.objects.filter(expense_date=day)

    sales_total = money_sum(sales, 'total_amount')
    gross_profit = money_sum(sales, 'gross_profit')
    expenses_total = money_sum(expenses, 'amount')
    cash_received = money_sum(sales, 'amount_paid')
    net_profit = gross_profit - expenses_total
    remaining_balance = cash_received - expenses_total

    # 2-month sales history (60 days)
    start_of_period = day - timedelta(days=59)
    sales_history = []
    current = start_of_period
    while current <= day:
        day_sales = Sale.objects.filter(sale_date=current)
        day_expenses = Expense.objects.filter(expense_date=current)
        day_sales_total = money_sum(day_sales, 'total_amount')
        day_profit = money_sum(day_sales, 'gross_profit')
        day_expense_total = money_sum(day_expenses, 'amount')
        day_cash = money_sum(day_sales, 'amount_paid')
        sales_history.append({
            'date': current,
            'sales_total': day_sales_total,
            'gross_profit': day_profit,
            'expenses_total': day_expense_total,
            'net_profit': day_profit - day_expense_total,
            'cash_received': day_cash,
        })
        current += timedelta(days=1)

    context = {
        'selected_date': day,
        'sales_total': sales_total,
        'gross_profit': gross_profit,
        'expenses_total': expenses_total,
        'cash_received': cash_received,
        'net_profit': net_profit,
        'remaining_balance': remaining_balance,
        'inventory_value': inventory_value(),
        'product_count': Product.objects.filter(is_active=True).count(),
        'service_count': Service.objects.filter(is_active=True).count(),
        'low_stock_products': Product.objects.filter(
            is_active=True,
            stock_quantity__lte=F('low_stock_threshold'),
        ).order_by('stock_quantity', 'name')[:8],
        'recent_sales': Sale.objects.prefetch_related('items').order_by('-created_at')[:8],
        'recent_expenses': Expense.objects.select_related('category').order_by('-created_at')[:8],
        'sales_history': sales_history,
    }
    return render(request, 'inventory/dashboard.html', context)


def product_list(request):
    products = Product.objects.select_related('category').order_by('name')
    query = request.GET.get('q', '').strip()
    stock_filter = request.GET.get('stock', '')
    if query:
        products = (
            products.filter(name__icontains=query)
            | products.filter(sku__icontains=query)
            | products.filter(brand__icontains=query)
        )
    if stock_filter == 'low':
        products = products.filter(stock_quantity__lte=F('low_stock_threshold'))
    elif stock_filter == 'active':
        products = products.filter(is_active=True)

    return render(request, 'inventory/product_list.html', {
        'products': products,
        'query': query,
        'stock_filter': stock_filter,
    })


def product_create(request):
    form = ProductForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Product saved.')
        return redirect('product_list')
    return render(request, 'inventory/product_form.html', {'form': form, 'title': 'New Product'})


def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, instance=product)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Product updated.')
        return redirect('product_list')
    return render(request, 'inventory/product_form.html', {'form': form, 'title': 'Edit Product'})


def service_list(request):
    services = Service.objects.select_related('category').order_by('name')
    query = request.GET.get('q', '').strip()
    if query:
        services = services.filter(name__icontains=query)
    return render(request, 'inventory/service_list.html', {'services': services, 'query': query})


def service_create(request):
    form = ServiceForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Service saved.')
        return redirect('service_list')
    return render(request, 'inventory/service_form.html', {'form': form, 'title': 'New Service'})


def service_update(request, pk):
    service = get_object_or_404(Service, pk=pk)
    form = ServiceForm(request.POST or None, instance=service)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Service updated.')
        return redirect('service_list')
    return render(request, 'inventory/service_form.html', {'form': form, 'title': 'Edit Service'})


def sale_list(request):
    sales = Sale.objects.prefetch_related('items').order_by('-sale_date', '-created_at')
    start = parse_date(request.GET.get('start', ''))
    end = parse_date(request.GET.get('end', ''))
    if start:
        sales = sales.filter(sale_date__gte=start)
    if end:
        sales = sales.filter(sale_date__lte=end)
    return render(request, 'inventory/sale_list.html', {'sales': sales, 'start': start, 'end': end})


def sale_detail(request, pk):
    sale = get_object_or_404(Sale.objects.prefetch_related('items__product', 'items__service'), pk=pk)
    return render(request, 'inventory/sale_detail.html', {'sale': sale})


def sale_create(request):
    form = SaleForm(request.POST or None)
    formset = SaleItemFormSet(request.POST or None)
    products_json = [
        {
            'id': product.id,
            'name': product.name,
            'price': str(product.selling_price),
            'cost': str(product.cost_price),
            'stock': product.stock_quantity,
        }
        for product in Product.objects.filter(is_active=True).order_by('name')
    ]
    services_json = [
        {
            'id': service.id,
            'name': service.name,
            'price': str(service.standard_price),
            'cost': str(service.delivery_cost),
        }
        for service in Service.objects.filter(is_active=True).order_by('name')
    ]

    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        line_forms = [line for line in formset if getattr(line, 'has_line_data', False)]
        if not line_forms:
            messages.error(request, 'Add at least one product or service line.')
        else:
            stock_errors = False
            for line in line_forms:
                item_type = line.cleaned_data['item_type']
                quantity = line.cleaned_data['quantity']
                product = line.cleaned_data.get('product')
                if item_type == SaleItem.ITEM_PRODUCT and product.stock_quantity < quantity:
                    line.add_error('quantity', f'Only {product.stock_quantity} in stock.')
                    stock_errors = True

            if not stock_errors:
                with transaction.atomic():
                    sale = form.save(commit=False)
                    paid_was_blank = form.cleaned_data.get('amount_paid') is None
                    if paid_was_blank:
                        sale.amount_paid = ZERO
                    sale.save()

                    for line in line_forms:
                        item_type = line.cleaned_data['item_type']
                        quantity = line.cleaned_data['quantity']
                        unit_price = line.cleaned_data.get('unit_price')
                        description = line.cleaned_data.get('description', '').strip()

                        if item_type == SaleItem.ITEM_PRODUCT:
                            product = Product.objects.select_for_update().get(pk=line.cleaned_data['product'].pk)
                            previous_quantity = product.stock_quantity
                            product.stock_quantity -= quantity
                            product.save(update_fields=['stock_quantity', 'updated_at'])
                            StockAdjustment.objects.create(
                                product=product,
                                adjustment_type=StockAdjustment.TYPE_SALE,
                                quantity_change=-quantity,
                                previous_quantity=previous_quantity,
                                new_quantity=product.stock_quantity,
                                note=f'Sold on {sale.invoice_number}',
                            )
                            SaleItem.objects.create(
                                sale=sale,
                                item_type=SaleItem.ITEM_PRODUCT,
                                product=product,
                                description=description or product.name,
                                quantity=quantity,
                                unit_price=unit_price if unit_price is not None else product.selling_price,
                            )
                        else:
                            service = line.cleaned_data['service']
                            SaleItem.objects.create(
                                sale=sale,
                                item_type=SaleItem.ITEM_SERVICE,
                                service=service,
                                description=description or service.name,
                                quantity=quantity,
                                unit_price=unit_price if unit_price is not None else service.standard_price,
                            )

                    sale.recalculate_totals()
                    if paid_was_blank:
                        sale.amount_paid = sale.total_amount
                        sale.recalculate_totals()

                messages.success(request, f'Sale {sale.invoice_number} saved.')
                return redirect(f'{reverse("sale_detail", args=[sale.pk])}?print=1')

    return render(request, 'inventory/sale_form.html', {
        'form': form,
        'formset': formset,
        'products_json': products_json,
        'services_json': services_json,
    })


def expense_list(request):
    expenses = Expense.objects.select_related('category').order_by('-expense_date', '-created_at')
    start = parse_date(request.GET.get('start', ''))
    end = parse_date(request.GET.get('end', ''))
    if start:
        expenses = expenses.filter(expense_date__gte=start)
    if end:
        expenses = expenses.filter(expense_date__lte=end)
    return render(request, 'inventory/expense_list.html', {'expenses': expenses, 'start': start, 'end': end})


def expense_create(request):
    form = ExpenseForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Expense saved.')
        return redirect('expense_list')
    return render(request, 'inventory/expense_form.html', {'form': form, 'title': 'New Expense'})


def expense_update(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    form = ExpenseForm(request.POST or None, instance=expense)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Expense updated.')
        return redirect('expense_list')
    return render(request, 'inventory/expense_form.html', {'form': form, 'title': 'Edit Expense'})


def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product_name = product.name
        product.delete()
        messages.success(request, f'Product "{product_name}" deleted.')
        return redirect('product_list')
    return render(request, 'inventory/confirm_delete.html', {'object': product, 'object_type': 'Product'})


def service_delete(request, pk):
    service = get_object_or_404(Service, pk=pk)
    if request.method == 'POST':
        service_name = service.name
        service.delete()
        messages.success(request, f'Service "{service_name}" deleted.')
        return redirect('service_list')
    return render(request, 'inventory/confirm_delete.html', {'object': service, 'object_type': 'Service'})


def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        expense_title = expense.title
        expense.delete()
        messages.success(request, f'Expense "{expense_title}" deleted.')
        return redirect('expense_list')
    return render(request, 'inventory/confirm_delete.html', {'object': expense, 'object_type': 'Expense'})


def sale_delete(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    if request.method == 'POST':
        invoice_number = sale.invoice_number
        sale.delete()
        messages.success(request, f'Sale {invoice_number} deleted.')
        return redirect('sale_list')
    return render(request, 'inventory/confirm_delete.html', {'object': sale, 'object_type': 'Sale'})


def product_category_delete(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk)
    # Check if category is in use
    if category.products.exists():
        messages.error(request, f'Cannot delete "{category.name}" - it has {category.products.count()} product(s) assigned to it.')
        return redirect('category_manager')
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Product category "{category_name}" deleted.')
        return redirect('category_manager')
    return render(request, 'inventory/confirm_delete.html', {'object': category, 'object_type': 'Product Category'})


def service_category_delete(request, pk):
    category = get_object_or_404(ServiceCategory, pk=pk)
    # Check if category is in use
    if category.services.exists():
        messages.error(request, f'Cannot delete "{category.name}" - it has {category.services.count()} service(s) assigned to it.')
        return redirect('category_manager')
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Service category "{category_name}" deleted.')
        return redirect('category_manager')
    return render(request, 'inventory/confirm_delete.html', {'object': category, 'object_type': 'Service Category'})


def expense_category_delete(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)
    # Check if category is in use
    if category.expenses.exists():
        messages.error(request, f'Cannot delete "{category.name}" - it has {category.expenses.count()} expense(s) assigned to it.')
        return redirect('category_manager')
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Expense category "{category_name}" deleted.')
        return redirect('category_manager')
    return render(request, 'inventory/confirm_delete.html', {'object': category, 'object_type': 'Expense Category'})


def stock_adjustment_create(request):
    form = StockAdjustmentForm(request.POST or None)
    adjustments = StockAdjustment.objects.select_related('product').order_by('-created_at')[:25]
    if request.method == 'POST' and form.is_valid():
        product = form.cleaned_data['product']
        adjustment_type = form.cleaned_data['adjustment_type']
        quantity = form.cleaned_data['quantity']
        note = form.cleaned_data['note']
        previous_quantity = product.stock_quantity

        if adjustment_type == StockAdjustment.TYPE_RESTOCK:
            new_quantity = previous_quantity + quantity
        elif adjustment_type == StockAdjustment.TYPE_REMOVE:
            new_quantity = previous_quantity - quantity
        else:
            new_quantity = quantity

        if new_quantity < 0:
            form.add_error('quantity', 'Stock cannot go below zero.')
        else:
            with transaction.atomic():
                product = Product.objects.select_for_update().get(pk=product.pk)
                previous_quantity = product.stock_quantity
                if adjustment_type == StockAdjustment.TYPE_RESTOCK:
                    new_quantity = previous_quantity + quantity
                elif adjustment_type == StockAdjustment.TYPE_REMOVE:
                    new_quantity = previous_quantity - quantity
                else:
                    new_quantity = quantity
                if new_quantity < 0:
                    form.add_error('quantity', 'Stock cannot go below zero.')
                else:
                    product.stock_quantity = new_quantity
                    product.save(update_fields=['stock_quantity', 'updated_at'])
                    StockAdjustment.objects.create(
                        product=product,
                        adjustment_type=adjustment_type,
                        quantity_change=new_quantity - previous_quantity,
                        previous_quantity=previous_quantity,
                        new_quantity=new_quantity,
                        note=note,
                    )
                    messages.success(request, 'Stock updated.')
                    return redirect('stock_adjustment_create')

    return render(request, 'inventory/stock_adjustment_form.html', {'form': form, 'adjustments': adjustments})


def category_manager(request):
    product_form = ProductCategoryForm(prefix='product')
    service_form = ServiceCategoryForm(prefix='service')
    expense_form = ExpenseCategoryForm(prefix='expense')

    if request.method == 'POST':
        category_type = request.POST.get('category_type')
        if category_type == 'product':
            product_form = ProductCategoryForm(request.POST, prefix='product')
            if product_form.is_valid():
                product_form.save()
                messages.success(request, 'Product category saved.')
                return redirect('category_manager')
        elif category_type == 'service':
            service_form = ServiceCategoryForm(request.POST, prefix='service')
            if service_form.is_valid():
                service_form.save()
                messages.success(request, 'Service category saved.')
                return redirect('category_manager')
        elif category_type == 'expense':
            expense_form = ExpenseCategoryForm(request.POST, prefix='expense')
            if expense_form.is_valid():
                expense_form.save()
                messages.success(request, 'Expense category saved.')
                return redirect('category_manager')

    return render(request, 'inventory/category_manager.html', {
        'product_form': product_form,
        'service_form': service_form,
        'expense_form': expense_form,
        'product_categories': ProductCategory.objects.order_by('name'),
        'service_categories': ServiceCategory.objects.order_by('name'),
        'expense_categories': ExpenseCategory.objects.order_by('name'),
    })


def report_context(start_date, end_date):
    sales = Sale.objects.filter(sale_date__range=[start_date, end_date])
    expenses = Expense.objects.filter(expense_date__range=[start_date, end_date])
    sales_total = money_sum(sales, 'total_amount')
    gross_profit = money_sum(sales, 'gross_profit')
    expenses_total = money_sum(expenses, 'amount')
    cash_received = money_sum(sales, 'amount_paid')

    daily_rows = []
    current = start_date
    while current <= end_date:
        day_sales = sales.filter(sale_date=current)
        day_expenses = expenses.filter(expense_date=current)
        day_sales_total = money_sum(day_sales, 'total_amount')
        day_profit = money_sum(day_sales, 'gross_profit')
        day_expense_total = money_sum(day_expenses, 'amount')
        day_cash = money_sum(day_sales, 'amount_paid')
        daily_rows.append({
            'date': current,
            'sales_total': day_sales_total,
            'gross_profit': day_profit,
            'expenses_total': day_expense_total,
            'net_profit': day_profit - day_expense_total,
            'cash_received': day_cash,
            'remaining_balance': day_cash - day_expense_total,
        })
        current += timedelta(days=1)

    top_items = (
        SaleItem.objects
        .filter(sale__sale_date__range=[start_date, end_date])
        .values('description', 'item_type')
        .annotate(quantity_sold=Sum('quantity'), sales_total=Sum('line_total'))
        .order_by('-sales_total')[:10]
    )

    return {
        'start_date': start_date,
        'end_date': end_date,
        'sales_total': sales_total,
        'gross_profit': gross_profit,
        'expenses_total': expenses_total,
        'cash_received': cash_received,
        'net_profit': gross_profit - expenses_total,
        'remaining_balance': cash_received - expenses_total,
        'inventory_value': inventory_value(),
        'daily_rows': daily_rows,
        'top_items': top_items,
        'sales': sales.order_by('-sale_date', '-created_at'),
        'expenses': expenses.order_by('-expense_date', '-created_at'),
    }


def reports(request):
    form, start_date, end_date = date_range_from_request(request)
    context = report_context(start_date, end_date)
    context['form'] = form
    return render(request, 'inventory/reports.html', context)


def append_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    header_fill = PatternFill('solid', fgColor='E2E8F0')
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        sheet.append(row)

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 38)
    return sheet


def export_xlsx(request):
    _, start_date, end_date = date_range_from_request(request)
    context = report_context(start_date, end_date)
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Summary'
    summary_rows = [
        ['Company', 'Fynnet Computer and Digital Services'],
        ['From', start_date.isoformat()],
        ['To', end_date.isoformat()],
        ['Sales', context['sales_total']],
        ['Gross Profit', context['gross_profit']],
        ['Expenses', context['expenses_total']],
        ['Net Profit', context['net_profit']],
        ['Cash Received', context['cash_received']],
        ['Remaining Balance', context['remaining_balance']],
        ['Current Inventory Value', context['inventory_value']],
    ]
    for row in summary_rows:
        summary.append(row)
    summary['A1'].font = Font(bold=True)
    summary.column_dimensions['A'].width = 28
    summary.column_dimensions['B'].width = 24

    append_sheet(
        workbook,
        'Daily Report',
        ['Date', 'Sales', 'Gross Profit', 'Expenses', 'Net Profit', 'Cash Received', 'Remaining Balance'],
        [
            [
                row['date'].isoformat(),
                row['sales_total'],
                row['gross_profit'],
                row['expenses_total'],
                row['net_profit'],
                row['cash_received'],
                row['remaining_balance'],
            ]
            for row in context['daily_rows']
        ],
    )
    append_sheet(
        workbook,
        'Sales',
        ['Invoice', 'Date', 'Customer', 'Phone', 'Payment', 'Status', 'Subtotal', 'Discount', 'Total', 'Cost', 'Profit', 'Paid', 'Balance Due'],
        [
            [
                sale.invoice_number,
                sale.sale_date.isoformat(),
                sale.customer_name,
                sale.customer_phone,
                sale.get_payment_method_display(),
                sale.get_status_display(),
                sale.subtotal,
                sale.discount_amount,
                sale.total_amount,
                sale.total_cost,
                sale.gross_profit,
                sale.amount_paid,
                sale.balance_due,
            ]
            for sale in context['sales']
        ],
    )
    append_sheet(
        workbook,
        'Sale Items',
        ['Invoice', 'Date', 'Type', 'Description', 'Quantity', 'Unit Price', 'Unit Cost', 'Line Total', 'Line Cost'],
        [
            [
                item.sale.invoice_number,
                item.sale.sale_date.isoformat(),
                item.get_item_type_display(),
                item.description,
                item.quantity,
                item.unit_price,
                item.unit_cost,
                item.line_total,
                item.line_cost,
            ]
            for item in SaleItem.objects.select_related('sale').filter(sale__sale_date__range=[start_date, end_date]).order_by('sale__sale_date')
        ],
    )
    append_sheet(
        workbook,
        'Expenses',
        ['Date', 'Category', 'Title', 'Amount', 'Payment Method', 'Notes'],
        [
            [
                expense.expense_date.isoformat(),
                expense.category.name if expense.category else '',
                expense.title,
                expense.amount,
                expense.payment_method,
                expense.notes,
            ]
            for expense in context['expenses']
        ],
    )
    append_sheet(
        workbook,
        'Products',
        ['SKU', 'Name', 'Category', 'Brand', 'Model', 'Cost Price', 'Selling Price', 'Stock', 'Low Stock Level', 'Inventory Value', 'Active'],
        [
            [
                product.sku,
                product.name,
                product.category.name if product.category else '',
                product.brand,
                product.model_number,
                product.cost_price,
                product.selling_price,
                product.stock_quantity,
                product.low_stock_threshold,
                product.inventory_value,
                'Yes' if product.is_active else 'No',
            ]
            for product in Product.objects.select_related('category').order_by('name')
        ],
    )
    append_sheet(
        workbook,
        'Services',
        ['Name', 'Category', 'Standard Price', 'Delivery Cost', 'Margin', 'Duration Minutes', 'Active'],
        [
            [
                service.name,
                service.category.name if service.category else '',
                service.standard_price,
                service.delivery_cost,
                service.service_margin,
                service.duration_minutes,
                'Yes' if service.is_active else 'No',
            ]
            for service in Service.objects.select_related('category').order_by('name')
        ],
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="fynnet-inventory-{start_date}-{end_date}.xlsx"'
    workbook.save(response)
    return response
